"""Экспериментальные инструменты для работы с Excel-файлами."""

from pathlib import Path

from openpyxl import load_workbook


def cell_to_text(value) -> str:
    """Преобразует значение Excel-ячейки в компактный текст."""
    if value is None:
        return ""

    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")

    return str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ")


def excel_to_text(path: Path) -> str:
    """Читает все листы книги и возвращает их содержимое в TSV-виде."""
    if not path.is_file():
        raise FileNotFoundError(f"Excel-файл не найден: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    parts = [f"===== ФАЙЛ: {path.name} ====="]

    try:
        for worksheet in workbook.worksheets:
            parts.append(f"--- ЛИСТ: {worksheet.title} ---")

            for row in worksheet.iter_rows(values_only=True):
                values = [cell_to_text(value) for value in row]

                while values and values[-1] == "":
                    values.pop()

                if values:
                    parts.append("\t".join(values))
    finally:
        workbook.close()

    return "\n".join(parts)
